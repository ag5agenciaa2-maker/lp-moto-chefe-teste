"""
fix_catalogo.py
Aplica todos os ajustes identificados na planilha catalogo_link360.xlsx
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import shutil, os, sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SRC  = 'assets/catalogo_link360.xlsx'
DEST = 'assets/catalogo_link360.xlsx'
BAK  = 'assets/catalogo_link360_BACKUP.xlsx'

# ── backup ────────────────────────────────────────────────────────────────────
shutil.copy2(SRC, BAK)
print(f'Backup criado: {BAK}')

# ── dados corrigidos (22 produtos únicos, sem duplicatas) ─────────────────────
# Colunas: id, nome, categoria, velocidade, potencia, autonomia,
#          descricao, img1, img2, img3, link, preco, preco_parc, badge, oferta, status

PRODUTOS = [
    # ── AUTOPROPELIDOS ────────────────────────────────────────────────────────
    (1,  'Giga',      'Autopropelidos', '32 km/h', '1000 W', '40 km',
     'Frente agressiva, farol e setas em LED e painel digital futurista. Escolha de quem quer sair do óbvio, chegar em silêncio e causar impacto.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/Giga_modelo.webp', '', '',
     'https://motochefebrasil.com.br/modelos/giga',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (2,  'JET MAX',   'Autopropelidos', '32 km/h', '1000 W', '55 km',
     'Visual esportivo de scooter robusta com tecnologia e conforto. Motor 1000W e maior autonomia da linha: 55 km por carga.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/JET-MAX-_-Oficial-2_b.webp', '', '',
     'https://motochefebrasil.com.br/modelos/jet-max',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (3,  'X12',       'Autopropelidos', '32 km/h', '1000 W', '40 km',
     'Scooter elétrica que redefine mobilidade urbana. Compacta, potente e tecnológica com 1000W e até 40 km de autonomia.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/x12b.webp', '', '',
     'https://motochefebrasil.com.br/modelos/x12/',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (4,  'BOB',       'Autopropelidos', '32 km/h', '1000 W', '40 km',
     'Posição confortável, plataforma baixa, assento amplo com espaço para carona e encosto traseiro. Ideal para o dia a dia urbano.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/bob_1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/bob',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (5,  'JET',       'Autopropelidos', '32 km/h', '1000 W', '40 km',
     'Design esportivo e bateria de lítio removível. Para quem quer estilo, agilidade e zero burocracia no trânsito.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/JET_1-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/jet',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (6,  'Joyzinha',  'Autopropelidos', '32 km/h', '600 W', '40 km',
     'Design compacto para o dia a dia. Parceira ideal para trabalho, faculdade ou passeio na orla — leve e econômica.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/5.webp', '', '',
     'https://motochefebrasil.com.br/modelos/joyzinha',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (7,  'Joy Classic','Autopropelidos','32 km/h', '600 W', '40 km',
     'Estilo clássico, conforto de sobra e autonomia suficiente para o seu dia. Silencioso, econômico e sem emplacamento.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/JOY-CLASSIC-_1-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/joy-classic',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (8,  'Joy Super', 'Autopropelidos', '32 km/h', '800 W', '40 km',
     'Motor 800W com tração firme e visual marcante. Mais potência que o clássico, mesma praticidade sem CNH.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/JOY-SUPER-1-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/joy-super',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (9,  'MC20 Mini', 'Autopropelidos', '32 km/h', '1000 W', '40 km',
     'Farol FULL LED, freio a disco hidráulico, bateria removível turbo 5A. Suporta até 180 kg — robusto de verdade.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/MC20-1-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/mc20-mini',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (10, 'MC21 Mini', 'Autopropelidos', '32 km/h', '1000 W', '40 km',
     'Equilíbrio perfeito entre design minimalista e desempenho elétrico. Bateria de lítio removível e carregamento em casa.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/mc21_mini_tin.webp', '', '',
     'https://motochefebrasil.com.br/modelos/mc21-mini',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (11, 'Mia',       'Autopropelidos', '32 km/h', '1000 W', '40 km',
     'Sem CNH, sem emplacamento. Resolução nº 996/2023 do Contran garante circulação legal. Motor 1000W e bateria removível.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/MIA-1-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/mia',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (12, 'Sofia',     'Autopropelidos', '32 km/h', '1000 W', '40 km',
     'Estilo clássico com tecnologia moderna. Motor 1000W, bateria removível e visual que agrada a todos os públicos.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/sofia_web3_tiny.webp', '', '',
     'https://motochefebrasil.com.br/modelos/sofia',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (13, 'Ret',       'Autopropelidos', '32 km/h', '1000 W', '40 km',
     'Estética retrô com motor 1000W e bateria de lítio. Presença de moto grande, praticidade de elétrica, sem burocracia.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/RET-1-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/ret',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    (14, 'Soma',      'Autopropelidos', '32 km/h', '1000 W', '40 km',
     'Sem CNH, sem emplacamento. Motor 1000W, bateria 60V 20Ah removível e 40 km de autonomia por carga.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/Soma-1-1-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/soma',
     '', '', 'Sem CNH', 'Não', 'Ativo'),

    # ── TRICICLOS (único registro com dupla categoria) ─────────────────────────
    (15, 'Mia Tri',   'Autopropelidos, Triciclos', '32 km/h', '800 W', '40 km',
     'Triciclo elétrico 800W para quem quer segurança extra, praticidade urbana e estilo moderno. Sem CNH.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/MIA-TRI-01-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/mia-tri',
     '', '', 'Sem CNH | Triciclo', 'Não', 'Ativo'),

    (16, 'Joy Tri',   'Autopropelidos, Triciclos', '30 km/h', '600 W', '40 km',
     'Chassi de três rodas com estabilidade total. Ideal para quem quer segurança sem abrir mão da liberdade elétrica.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/JOY-TRI-1-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/joy-tri',
     '', '', 'Sem CNH | Triciclo', 'Não', 'Ativo'),

    (17, 'BIG TRI',   'Autopropelidos, Triciclos', '32 km/h', '1000 W', '40 km',
     'Liberdade com segurança. Design retrô, conforto ampliado e tecnologia elétrica moderna. Motor 1000W, sem CNH.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/big_tri.webp', '', '',
     'https://motochefebrasil.com.br/modelos/big-tri/',
     '', '', 'Sem CNH | Triciclo', 'Não', 'Ativo'),

    (18, 'VED',       'Autopropelidos, Triciclos', '32 km/h', '1000 W', '40 km',
     'Mobilidade elétrica acessível e confortável. Motor 1000W, chassi de três rodas e autonomia de até 40 km. Sem CNH.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/ved_1-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/ved',
     '', '', 'Sem CNH | Triciclo', 'Não', 'Ativo'),

    # ── CICLOMOTORES (requer CNH categoria A ou B) ─────────────────────────────
    (19, 'X11',       'Ciclomotor', '50 km/h', '2000 W', '80 km',
     'Design robusto e presença marcante. Motor 2000W, até 80 km de autonomia — o mais completo da categoria ciclomotor.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/x11__-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/x11',
     '', '', 'Maior Autonomia', 'Sim', 'Ativo'),

    (20, 'MC20',      'Ciclomotor', '50 km/h', '2000 W', '40 km',
     '2000W ou 3000W, NFC, alarme com bloqueio, painel digital e até 80 km com bateria extra. Visual chopper, IP65.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/mc20_pop-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/mc20/',
     '', '', 'NFC + Alarme', 'Sim', 'Ativo'),

    (21, 'X15',       'Ciclomotor, Triciclos', '50 km/h', '3000 W', '40 km',
     'Triciclo elétrico potente. Motor 3000W para aceleração firme e desempenho consistente em qualquer pista.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/x15.webp', '', '',
     'https://motochefebrasil.com.br/modelos/x15',
     '', '', 'Mais Potente', 'Não', 'Ativo'),

    (22, 'Roma',      'Ciclomotor', '50 km/h', '3000 W', '50 km',
     'Charme retrô europeu com tecnologia elétrica moderna. Motor 3000W, 50 km de autonomia e visual inconfundível.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/roma_ai_sombra-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/roma',
     '', '', 'Design Premium', 'Sim', 'Ativo'),

    # ── E-BIKES ────────────────────────────────────────────────────────────────
    (23, 'GRID',      'E-Bikes', '32 km/h', '750 W', '35 km',
     'E-bike urbana ágil, robusta e divertida de pilotar. Motor 750W, freio a disco e bateria removível.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/10/GRID-LATERAL1-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/grid/',
     '', '', 'Pedal Assistido', 'Não', 'Ativo'),

    (24, 'Style',     'E-Bikes', '32 km/h', '750 W', '35 km',
     'E-bike urbana 750W (pico 1000W), bateria 48V 15,6Ah removível, freio a disco hidráulico — assinada por Diego Ribas.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/style-1-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/style',
     '', '', 'Edição Diego Ribas', 'Sim', 'Ativo'),

    (25, 'Liberty',   'E-Bikes', '32 km/h', '500 W', '35 km',
     'Motor 500W (pico 800W), bateria 48V 13Ah removível e modos por aceleração ou pedal assistido PAS 5 níveis.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/liberty_product_web.webp', '', '',
     'https://motochefebrasil.com.br/modelos/liberty',
     '', '', 'Pedal Assistido', 'Não', 'Ativo'),

    (26, 'SPACE',     'E-Bikes', '32 km/h', '750 W', '35 km',
     'Motor 750W, bateria 48V 12Ah removível, 7 marchas. Aceleração no punho ou pedal assistido PAS 5 níveis.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/space_ia-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/space',
     '', '', 'Pedal Assistido', 'Não', 'Ativo'),

    (27, 'Retrô',     'E-Bikes', '32 km/h', '500 W', '35 km',
     'Visual de bike urbana clássica com praticidade de e-bike moderna. Motor 500W, pedal assistido e bateria removível.',
     'https://motochefebrasil.com.br/wp-content/uploads/2025/11/retro_hero-1.webp', '', '',
     'https://motochefebrasil.com.br/modelos/retro',
     '', '', 'Pedal Assistido', 'Não', 'Ativo'),
]

# ── abre e recria a planilha ─────────────────────────────────────────────────
wb  = openpyxl.load_workbook(SRC)
ws  = wb.active

# apaga conteúdo existente a partir da linha 3 (mantém cabeçalhos)
for row in ws.iter_rows(min_row=3):
    for cell in row:
        cell.value = None

# ── estilos ────────────────────────────────────────────────────────────────────
GOLD        = 'FFC9A227'
DARK        = 'FF1A1F3D'
WHITE       = 'FFFFFFFF'
LIGHT_GRAY  = 'FFF5F5F5'
GREEN_SOFT  = 'FFE8F5E9'
GREEN_TEXT  = 'FF2E7D32'
ORANGE_SOFT = 'FFFFF3E0'
ORANGE_TEXT = 'FFE65100'

hdr_fill  = PatternFill('solid', fgColor=DARK)
hdr_font  = Font(bold=True, color=WHITE, size=9)
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

alt_fill  = PatternFill('solid', fgColor=LIGHT_GRAY)
gold_fill = PatternFill('solid', fgColor='FFFFF8E7')
thin      = Side(style='thin', color='DDDDDD')
border    = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── atualiza cabeçalhos (adiciona 2 colunas novas) ────────────────────────────
# Colunas originais (linha 2): ID, Nome, Categoria, Vel, Pot, Auto,
#                              Desc, Img1, Img2, Img3, Link, Preço,
#                              Oferta, Status
# Novas: Preço Parcelado (col 13), Badge (col 14) — Oferta vira 15, Status 16

headers = [
    'ID',
    'Nome do Produto *',
    'Categoria *',
    'Velocidade Máx.',
    'Potência (Motor)',
    'Autonomia',
    'Descrição *',
    '🖼️ Imagem Destaque (URL)',
    'Imagem 2 (URL)',
    'Imagem 3 (URL)',
    'Link do Produto (URL)',
    'Preço (R$)',
    'Preço Parcelado',
    'Badge / Tag',
    '🔥 Oferta em Destaque?',
    'Status',
]

for col_idx, hdr in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=hdr)
    cell.fill    = hdr_fill
    cell.font    = hdr_font
    cell.alignment = hdr_align
    cell.border  = border

# ── preenche dados ─────────────────────────────────────────────────────────────
for r_idx, p in enumerate(PRODUTOS, start=3):
    (pid, nome, cat, vel, pot, aut,
     desc, img1, img2, img3, link,
     preco, preco_parc, badge, oferta, status) = p

    values = [pid, nome, cat, vel, pot, aut,
              desc, img1, img2, img3, link,
              preco, preco_parc, badge, oferta, status]

    is_even   = (r_idx % 2 == 0)
    is_oferta = (oferta == 'Sim')

    for c_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border    = border
        cell.alignment = Alignment(vertical='center', wrap_text=(c_idx == 7))

        if is_oferta:
            cell.fill = gold_fill
        elif is_even:
            cell.fill = alt_fill

    # coluna Oferta: destaque visual
    oferta_cell = ws.cell(row=r_idx, column=15)
    if oferta == 'Sim':
        oferta_cell.font = Font(bold=True, color=ORANGE_TEXT)
        oferta_cell.fill = PatternFill('solid', fgColor=ORANGE_SOFT)
    else:
        oferta_cell.font = Font(color='FF888888')

    # coluna Status
    status_cell = ws.cell(row=r_idx, column=16)
    if status == 'Ativo':
        status_cell.font = Font(bold=True, color=GREEN_TEXT)
        status_cell.fill = PatternFill('solid', fgColor=GREEN_SOFT)

# ── larguras das colunas ──────────────────────────────────────────────────────
col_widths = [5, 18, 26, 14, 16, 12, 60, 55, 45, 45, 45, 14, 18, 22, 18, 10]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── altura das linhas de dados ────────────────────────────────────────────────
for r in range(3, 3 + len(PRODUTOS)):
    ws.row_dimensions[r].height = 52

# ── salva ──────────────────────────────────────────────────────────────────────
wb.save(DEST)
print(f'Planilha salva: {DEST}')
print(f'Produtos unicos: {len(PRODUTOS)} (era 32 com duplicatas)')
print('Ajustes aplicados:')
print('  - Duplicatas removidas (18 → 27 produtos únicos, triciclos com dupla categoria)')
print('  - Coluna "Preço Parcelado" adicionada')
print('  - Coluna "Badge / Tag" adicionada e preenchida')
print('  - Descrições repetidas corrigidas (JET, Joy Super, Ret, Retrô, X12)')
print('  - Bug de URL do X15 Triciclos corrigido')
print('  - Destaques marcados: X11, MC20, Roma, Style (Sim)')
print('  - Formatação visual aplicada (zebra, badges, ouro nos destaques)')
